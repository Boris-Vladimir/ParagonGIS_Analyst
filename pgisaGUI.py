#==============================================================================
# IMPORTS
#==============================================================================
import os
import sqlite3
import xlrd
from datetime import datetime, date
from openpyxl import Workbook
from jdcal import gcal2jd, jd2gcal, MJD_0  # for the compiler :(
from time import sleep
from tkinter import HORIZONTAL, BOTTOM, CENTER, FALSE, ALL, E, W, EW, NS, \
    SEPARATOR, Tk, Frame, Label, Button, Message, Menu, Canvas, Toplevel, \
    StringVar, BooleanVar, filedialog, ttk, SUNKEN, TOP, TRUE, BOTH, LEFT, \
    Text, NORMAL, DISABLED, Entry, simpledialog, DoubleVar, IntVar, \
    Checkbutton, VERTICAL, RIGHT, Radiobutton, messagebox, font as fontTK


class window(object):

    def __init__(self):
        global icon_color
        global icon_scale
        global cgi_time
        self.master = Tk()  # Tk() object
        self.master.title('Paragon GIS Analyst - ver. 1.5')  # window name
        icons = os.getcwd() + os.sep + "icons" + os.sep  # path to icons
        icon = icons + "maps.ico"
        self.master.iconbitmap(icon)  # window icon
        self.master.resizable(width=FALSE, height=FALSE)
        self.master.geometry("520x120")
        self.file_name = ""  # the name of the EXEL file
        self.last_dir = "C:/"

        # user theme
        self.style = ttk.Style()
        self.style.theme_create( "vladimir", parent="alt", settings={
            "TNotebook": {"configure": {"tabmargins": [0, 0, 0, 0] , 
                                        "background": "#404040"}},
            "TNotebook.Tab": {
                "configure": {"padding": [1, 1], "background": "#404040" ,
                              "foreground": '#ff8c00'},
                "map": {"background": [("selected", '#ff8c00')],
                        "foreground": [("selected", '#404040')],}}})

        # to use in frame, message, labels and buttons ------------------------
        self.message = StringVar()
        self.message.set("\nSelecciona um ficheiro EXCEL")
        bg = "gray25"
        bg1 = "dark orange"
        fc = "white smoke"
        font = ("Helvetica", "8", "bold")
        text0 = " ----- "
        text1 = " Boris & Vladimir Software "
        text = text0 + text1 + text0
        self.cor_icon = StringVar()
        self.scale_icon = DoubleVar()
        self.cgi_time = IntVar()

        # Menu ----------------------------------------------------------------
        self.menu = Menu(self.master)
        self.master.config(menu=self.menu)
        filemenu = Menu(self.menu)
        self.menu.add_cascade(label="Ficheiro", menu=filemenu)
        filemenu.add_command(label="Abrir...", command=self.__callback)
        filemenu.add_command(label="Gravar", command=self.__callback_2)
        filemenu.add_command(label="Sair", command=self.__callback_3)
        
        bdmenu = Menu(self.menu)
        self.menu.add_cascade(label="BD Operadoras Móveis", menu=bdmenu)
        bdmenu.add_command(label="Actualizar", command=None)  # FALTA COMANDO ###############################################################################
        #bdmenu.add_command(label="Consultar", command=lambda: (DbGui(self.master)))  # FALTA COMANDO ###############################################################################
        bdmenu.add_command(label="Consultar", command=lambda:(self.__queryDB()))
        exportmenu = Menu(bdmenu)
        bdmenu.add_cascade(label="Exportar", menu=exportmenu)
        exportmenu.add_command(label="MEO", command=None)  # FALTA COMANDO ###############################################################################
        exportmenu.add_command(label="NOS", command=None)  # FALTA COMANDO ###############################################################################
        exportmenu.add_command(label="Vodafone", command=None)  # FALTA COMANDO ###############################################################################
        exportmenu.add_command(label="Todos", command=None)  # FALTA COMANDO ###############################################################################

        docsmenu = Menu(self.menu)
        docs = ["docs\manual.pdf", "docs\icons.pdf", "docs\colors.pdf"]
        self.menu.add_cascade(label="Documentação", menu=docsmenu)
        docsmenu.add_command(label="Manual",
                             command=lambda: (self.__open_file(docs[0])))
        docsmenu.add_command(label="Ícones",
                             command=lambda: (self.__open_file(docs[1])))
        docsmenu.add_command(label="Cores",
                             command=lambda: (self.__open_file(docs[2])))

        helpmenu = Menu(self.menu)
        self.menu.add_cascade(label='Ajuda', menu=helpmenu)
        helpmenu.add_command(label="Sobre", command=self.__about)
        helpmenu.add_command(label="Ver erros",
                             command=lambda: (self.__open_file("erros.log")))

        # Frame to suport butons, labels and separators -----------------------
        self.f = Frame(self.master, bg=bg)
        self.f.pack_propagate(0)  # don't shrink
        self.f.pack(side=BOTTOM, padx=0, pady=0)

        # Message, Labels and Entries -----------------------------------------
        self.l1 = Message(
            self.f, bg=bg1, bd=5, fg=bg, textvariable=self.message,
            font=("Helvetica", "13", "bold italic"), width=500).grid(
            row=0, columnspan=6, sticky=EW, padx=5, pady=5)

        self.l6 = Label(
            self.f, text=text, font=("Helvetica", "11", "bold"), bg=bg, fg=bg1
            ).grid(row=3, column=2, columnspan=3, sticky=EW, pady=5)

        # Buttons -------------------------------------------------------------
        self.b0 = Button(
            self.f, text="Abrir EXCEL...", command=self.__callback, width=10,
            bg="forest green", fg=fc, font=font
            ).grid(row=3, column=0, padx=5, sticky=W)
        self.b1 = Button(
            self.f, text="Gravar KMZ", command=self.__callback_2, width=10,
            bg="DodgerBlue4", fg=fc, font=font
            ).grid(row=3, column=1, sticky=W)
        self.b2 = Button(
            self.f, text="Sair", command=self.__callback_3, width=10,
            bg="orange red", fg=fc, font=font
            ).grid(row=3, column=5, sticky=E, padx=5)

        # Mainloop ------------------------------------------------------------
        self.master.mainloop()

    def __callback(self):  # "Abrir ECXEL..." button handler ------------------
        '''
        None -> None

        Opens a new window (filedialog.askopenfilename) to choose the
        EXCEL file that is necessary to make the KMZ file.
        '''
        title = 'Selecciona um ficheiro Excel'
        message = 'Ficheiro EXCEL carregado em memória.\nTransforma-o em KMZ!'
        self.file_name = filedialog.askopenfilename(title=title,
                                                    initialdir=self.last_dir)
        self.last_dir = self.file_name[:self.file_name.rfind('/')]

        if self.file_name[self.file_name.rfind('.') + 1:] != 'xls' and \
                self.file_name[self.file_name.rfind('.') + 1:] != 'xlsx':
            message = self.file_name + ' não é um ficheiro Excel válido!'
        else:
            sleep(1)
            message = 'Ficheiro EXCEL carregado em memória.\n\
            Defina as propriedades do Icon'
            self.message.set(message)
            IconDialog(self.master)
            self.cor_icon.set(icon_color)
            self.scale_icon.set(icon_scale)
            self.cgi_time.set(cgi_time)
            sleep(1)
            message = 'Propriedades do Icon definidas.\nClique em Gravar'
            self.message.set(message)

        self.message.set(message)

    def __callback_2(self):  # "Gravar KMZ" button handler --------------------
        '''
        None -> None

        Calls the function self.__threat()
        '''
        sleep(1)
        message = 'Propriedades do Icon definidas.\nClique em Gravar'
        if self.message.get() != message:
            self.message.set("\nEscolhe um ficheiro EXCEL primeiro")
            self.master.update_idletasks()
        else:
            '''
            self.message.set("\nA processar...")
            '''
            ui = Controler()  # start interface
            ui.input_file_name = self.file_name
            filter_list = [self.cgi_time.get(), 'mvn', '234', 'n', None]
            ui.filter_list = filter_list
            # choose build an icon or polygon file ----------------------------
            color = self.cor_icon.get()
            scale = self.scale_icon.get()
            kmz = ui.icon_or_polygon(color, scale)
            kmz.file_name = ui.input_file_name
            # make the new Excel file -----------------------------------------
            kmz_file = BusinessLogic()
            if isinstance(kmz, Icon):
                kmz_file.build_icon_file(kmz, filter_list)
            else:
                kmz_file.build_polygon_file(kmz, filter_list)

            self.message.set('\n Ficheiro criado com sucesso')

            self.master.update_idletasks()
            sleep(1)

    def __callback_3(self):  # "Sair" button handler --------------------------
        '''
        None -> None

        Kills the window
        '''
        self.master.destroy()

    def __about(self):
        '''
        None -> None

        Associated with the Help Menu.
        Creates a new window with the "About" information
        '''
        appversion = "1.5"
        appname = "Paragon GIS Analyst"
        copyright = 14 * ' ' + '(c) 2014' + 12 * ' ' + \
            'SDATO - DP - UAF - GNR\n' + 34 * ' '\
            + "No Rights Reserved... ask the code ;)"
        licence = 18 * ' ' + 'http://opensource.org/licenses/GPL-3.0\n'
        contactname = "Nuno Venâncio"
        contactphone = "(00351) 969 564 906"
        contactemail = "venancio.gnr@gmail.com"

        message = "Version: " + appversion + 5 * "\n"
        message0 = "Copyleft: " + copyright + "\n" + "Licença: " + licence
        message1 = contactname + '\n' + contactphone + '\n' + contactemail

        icons = os.getcwd() + os.sep + "icons" + os.sep  # path to icons
        icon = icons + "maps.ico"

        tl = Toplevel(self.master)
        tl.configure(borderwidth=5)
        tl.title("Sobre...")
        tl.iconbitmap(icon)
        tl.resizable(width=FALSE, height=FALSE)
        f1 = Frame(tl, borderwidth=2, relief=SUNKEN, bg="gray25")
        f1.pack(side=TOP, expand=TRUE, fill=BOTH)

        l0 = Label(f1, text=appname, fg="white", bg="gray25",
                   font=('courier', 16, 'bold'))
        l0.grid(row=0, column=0, sticky=W, padx=10, pady=5)
        l1 = Label(f1, text=message, justify=CENTER,
                   fg="white", bg="gray25")
        l1.grid(row=2, column=0, sticky=E, columnspan=3, padx=10, pady=0)
        l2 = Label(f1, text=message0,
                   justify=LEFT, fg="white", bg="gray25")
        l2.grid(row=6, column=0, columnspan=2, sticky=W, padx=10, pady=0)
        l3 = Label(f1, text=message1,
                   justify=CENTER, fg="white", bg="gray25")
        l3.grid(row=7, column=0, columnspan=2, padx=10, pady=0)

        button = Button(tl, text="Ok", command=tl.destroy, width=10)
        button.pack(pady=5)

    def __queryDB(self):
        y = datetime.now().year
        m = datetime.now().month
        d = datetime.now().day
        today = "{:0>2d}".format(d) + '/' + "{:0>2d}".format(m) + '/' + str(y)
        self.query_dict = {'cgi': None,
                           'tec': '234',
                           'date': {'first': '01/01/2013', 'last': today, 
                                    'd_all': False},
                           'pol': {'cp': None, 'loc': None, 'con': None, 
                                   'dis': None},
                           'fis': {'form': None, 'lat': None, 'lon': None, 
                                   'lat2': None, 'lon2': None, 'radius': None},
                           'kmz': {'ico_or_pol': None, 'n': 338, 'scale': 1,
                                   'color': "orange", 'radius': None,
                                   'alt': 35, 'amp': 110}}
        # MAIN WINDOW #########################################################
        # --- Theme configuration ---------------------------------------------
        bg = "gray25"
        fg = "dark orange"
        font = ("Helvetica", "10", "bold")
        self.style.theme_use("vladimir")
        tfs = ttk.Style()
        tfs.configure("TFrame", background="#404040")
        # --- Toplevel --------------------------------------------------------
        tl = Toplevel(self.master)
        #tl.geometry("410x420")
        tl.geometry("530x420")
        tl.configure(borderwidth=5, bg=bg)
        tl.title("Pesquisar Base de Dados")
        icons = os.getcwd() + os.sep + "icons" + os.sep  # path to icons
        icon = icons + "maps.ico"
        tl.iconbitmap(icon)
        tl.resizable(width=FALSE, height=FALSE)
        # --- Base Frame ------------------------------------------------------
        f1 = Frame(tl, bg=bg)
        f1.pack_propagate(0)  # don't shrink
        f1.pack(side=BOTTOM, padx=0, pady=0, expand=TRUE, fill=BOTH)
        # --- Notebook --------------------------------------------------------
        nf = fontTK.Font(root=tl, family='helvetica', size='12', weight='bold')
        s = ttk.Style()
        s.configure('.', font=nf)
        n = ttk.Notebook(f1)
        n.pack(side=TOP)        
        # --- Messages / Log Frame --------------------------------------------
        self.msg = StringVar()
        self.msg.set("Log...\nMensagens do programa...")
        l1 = Message(f1, bg=fg, bd=5, fg=bg, textvariable=self.msg,
            font=("Helvetica", "10", "bold italic"), width=500)
        l1.pack_propagate(0)
        l1.pack(expand=TRUE, fill=BOTH, pady=5)
        # --- Pesquisar Button ------------------------------------------------
        b = Button(f1, text="Pesquisar", command=self.__queryDB_button,
                   width=10, bg="forest green", fg="white smoke", 
                   font=("Helvetica", "8", "bold"))
        b.pack(side=BOTTOM, anchor=E)

        # CGI TAB #############################################################
        # --- Main Frame ------------------------------------------------------
        tab_cgi = ttk.Frame(n)
        n.add(tab_cgi, text='CGI')
        # --- inside Frame ----------------------------------------------------
        f_cgi = ttk.Frame(tab_cgi)
        f_cgi.pack_propagate(0)
        f_cgi.pack(padx=16, pady=33)
        # --- Labels ----------------------------------------------------------
        cgi_l1 = Label(f_cgi, text=None, bg=bg)        
        cgi_l2 = Label(f_cgi, text="Operadora", bg=bg, fg=fg, width=7,
                       justify=RIGHT)        
        cgi_l3 = Label(f_cgi, text="LAC", bg=bg, fg=fg, width=7, justify=RIGHT)        
        cgi_l4 = Label(f_cgi, text="CID", bg=bg, fg=fg, width=7, justify=RIGHT)        
        cgi_l5 = Label(f_cgi, text="CGI", bg=bg, fg=fg)
        cgi_l1.grid(row=0, pady=0)
        cgi_l2.grid(row=1, column=0, padx=5, pady=5)
        cgi_l3.grid(row=1, column=2, pady=5)
        cgi_l4.grid(row=1, column=4, pady=5)
        cgi_l5.grid(row=4, pady=5, padx=5)
        # --- Entries ---------------------------------------------------------
        self.cgi_e1 = Entry(f_cgi, width=2)
        self.cgi_e2 = Entry(f_cgi, width=4)
        self.cgi_e3 = Entry(f_cgi, width=6)
        self.cgi_e4 = Entry(f_cgi, width=25)
        self.cgi_e1.grid(row=1, column=1, pady=5)
        self.cgi_e2.grid(row=1, column=3, pady=5)
        self.cgi_e3.grid(row=1, column=5, pady=5)
        self.cgi_e4.grid(row=4, column=1, columnspan=4, pady=5)
        # --- Separator -------------------------------------------------------
        cgi_s = ttk.Separator(f_cgi, orient=HORIZONTAL)
        cgi_s.grid(row=3, column=0, columnspan=6, sticky=EW, padx=5, pady=5)        

        # TECNOLOGIA TAB ######################################################
        # --- Main Frame ------------------------------------------------------
        tab_tec = ttk.Frame(n)
        n.add(tab_tec, text='Tecnologia')
        # --- inside Frame ----------------------------------------------------
        f_tec = ttk.Frame(tab_tec)
        f_tec.pack_propagate(0)
        f_tec.pack(padx=40, pady=45)
        # --- variables -------------------------------------------------------
        self.tec_2g = IntVar()
        self.tec_3g = IntVar()
        self.tec_4g = IntVar()
        # --- Checkbuttons ----------------------------------------------------
        tec_cb1 = Checkbutton(f_tec, text="2G", onvalue=2,
                              variable=self.tec_2g, bg=bg, fg=fg)
        tec_cb2 = Checkbutton(f_tec, text="3G", onvalue=3,
                              variable=self.tec_3g, bg=bg, fg=fg)
        tec_cb3 = Checkbutton(f_tec, text="4G", onvalue=4,
                              variable=self.tec_4g, bg=bg, fg=fg)  
        tec_cb1.grid(row=1, pady=5, padx=5)
        tec_cb2.grid(row=2, pady=5, padx=5)
        tec_cb3.grid(row=3, pady=5, padx=5)
        # --- Labels ----------------------------------------------------------
        tec_l1 = Label(f_tec, text="(GSM)", bg=bg, fg=fg)        
        tec_l2 = Label(f_tec, text="(WCDMA, HSPA, UMTS)", bg=bg, fg=fg)        
        tec_l3 = Label(f_tec, text="(WIMAX, LTE)", bg=bg, fg=fg)
        tec_l1.grid(row=1, column=1, pady=5, sticky=W)
        tec_l2.grid(row=2, column=1, pady=5)
        tec_l3.grid(row=3, column=1, pady=5, sticky=W)

        # DATA TAB ############################################################
        # --- Main Frame ------------------------------------------------------
        tab_dat = ttk.Frame(n)
        n.add(tab_dat, text='Data')
        # --- inside Frame ----------------------------------------------------
        f_dat = ttk.Frame(tab_dat)
        f_dat.pack_propagate(0)
        f_dat.pack(padx=0, pady=55)
        # --- variables -------------------------------------------------------
        days = ["{:0>2d}".format(n + 1) for n in range(31)]
        months = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET',
                'OUT', 'NOV', 'DEC']
        years = [n + 1 for n in range(2012, datetime.now().year)]
        self.date_all = IntVar()
        # --- Labels ----------------------------------------------------------
        dat_l1 = Label(f_dat, text="De", bg=bg, fg=fg)        
        dat_l2 = Label(f_dat, text="/", bg=bg, fg=fg)        
        dat_l3 = Label(f_dat, text="/", bg=bg, fg=fg)        
        dat_l4 = Label(f_dat, text="a", bg=bg, fg=fg)        
        dat_l5 = Label(f_dat, text="/", bg=bg, fg=fg)        
        dat_l6 = Label(f_dat, text="/", bg=bg, fg=fg)        
        dat_l7 = Label(f_dat, text=" ", bg=bg, fg=fg)
        dat_l1.grid(row=1, column=0, pady=5, padx=5)
        dat_l2.grid(row=1, column=2, pady=5, padx=0)
        dat_l3.grid(row=1, column=4, pady=5, padx=0)
        dat_l4.grid(row=1, column=6, pady=5, padx=2)
        dat_l5.grid(row=1, column=8, pady=5, padx=0)
        dat_l6.grid(row=1, column=10, pady=5, padx=0)
        dat_l7.grid(row=1, column=12, pady=5, padx=0)
        # --- Comboboxes ------------------------------------------------------
        self.dat_cb1 = ttk.Combobox(f_dat, width=3, textvariable=None,
                                    values=days)        
        self.dat_cb2 = ttk.Combobox(f_dat, width=5, textvariable=None,
                                    values=months)        
        self.dat_cb3 = ttk.Combobox(f_dat, width=4, textvariable=None,
                                    values=years)                
        self.dat_cb4 = ttk.Combobox(f_dat, width=3, textvariable=None,
                                    values=days)               
        self.dat_cb5 = ttk.Combobox(f_dat, width=5, textvariable=None,
                                    values=months)                
        self.dat_cb6 = ttk.Combobox(f_dat, width=4, textvariable=None,
                                    values=years)
        self.dat_cb1.grid(row=1, column=1, padx=0)
        self.dat_cb2.grid(row=1, column=3, padx=0)
        self.dat_cb3.grid(row=1, column=5, padx=0)
        self.dat_cb4.grid(row=1, column=7, padx=0)
        self.dat_cb5.grid(row=1, column=9, padx=0)
        self.dat_cb6.grid(row=1, column=11, padx=0)
        # --- Separator -------------------------------------------------------
        dat_s = ttk.Separator(f_dat, orient=HORIZONTAL)
        dat_s.grid(row=2, columnspan=13, sticky=EW, pady=5, padx=5)
        # --- Radiobuttons ----------------------------------------------------
        dat_rb1 = Radiobutton(f_dat, text='Todas', variable=self.date_all, 
                              value=1, bg=bg, fg=fg)
        dat_rb1.grid(row=3, column=0, columnspan=6, pady=5, padx=5)
        dat_rb2 = Radiobutton(f_dat, text='Mais Recentes', 
                              variable=self.date_all, value=0, bg=bg, fg=fg)        
        dat_rb2.grid(row=3, column=6, columnspan=5, pady=5, padx=5)
        dat_rb1.deselect()
        dat_rb2.select()

        # LOCALIZAÇÃO ADMINISTRATIVA TAB ######################################
        # --- Main Frame ------------------------------------------------------
        tab_pol = ttk.Frame(n)
        n.add(tab_pol, text='Localização Administrativa')
        # --- inside Frame ----------------------------------------------------
        f_pol = ttk.Frame(tab_pol)
        f_pol.pack_propagate(0)
        f_pol.pack(padx=0, pady=40)
        # --- Labels ----------------------------------------------------------
        pol_l1 = Label(f_pol, text="Cod Postal", bg=bg, fg=fg)        
        pol_l2 = Label(f_pol, text="Localidade", bg=bg, fg=fg)        
        pol_l3 = Label(f_pol, text="Concelho", bg=bg, fg=fg)        
        pol_l4 = Label(f_pol, text="Distrito", bg=bg, fg=fg)
        pol_l1.grid(row=2, column=0, padx=5, pady=5)
        pol_l2.grid(row=3, column=0, padx=5, pady=5)
        pol_l3.grid(row=4, column=0, padx=5, pady=5)
        pol_l4.grid(row=5, column=0, padx=5, pady=5)
        # --- Entries ---------------------------------------------------------
        self.pol_e1 = Entry(f_pol)        
        self.pol_e2 = Entry(f_pol)        
        self.pol_e3 = Entry(f_pol)        
        self.pol_e4 = Entry(f_pol)
        self.pol_e1.grid(row=2, column=1)
        self.pol_e2.grid(row=3, column=1)
        self.pol_e3.grid(row=4, column=1)
        self.pol_e4.grid(row=5, column=1)

        # LOCALIZAÃÇÃO FÍSICA TAB #############################################
        # --- Main Frame ------------------------------------------------------
        tab_fis = ttk.Frame(n)
        n.add(tab_fis, text='Localização Fí­sica')
        # --- inside Frame ----------------------------------------------------
        f_fis = ttk.Frame(tab_fis)
        f_fis.pack_propagate(0)
        f_fis.pack(padx=5, pady=10)
        # --- Labels ----------------------------------------------------------
        fis_l1 = Label(f_fis, text="Cí­rculo", bg=bg, fg=fg, font=font)        
        fis_l2 = Label(f_fis, text="Latitude", bg=bg, fg=fg)        
        fis_l3 = Label(f_fis, text="Longitude", bg=bg, fg=fg)        
        fis_l4 = Label(f_fis, text="Raio (Kms)", bg=bg, fg=fg)        
        fis_l5 = Label(f_fis, text="Quadrado", bg=bg, fg=fg, font=font)        
        fis_l6 = Label(f_fis, text="Latitude 1", bg=bg, fg=fg)        
        fis_l7 = Label(f_fis, text="Longitude 1", bg=bg, fg=fg)        
        fis_l8 = Label(f_fis, text="Latitude 2", bg=bg, fg=fg)        
        fis_l9 = Label(f_fis, text="Longitude 2", bg=bg, fg=fg)
        fis_l1.grid(row=1, column=0, columnspan=4, padx=5, pady=0)
        fis_l2.grid(row=2, column=0, padx=5, pady=5)
        fis_l3.grid(row=2, column=2, padx=5, pady=5)
        fis_l4.grid(row=3, column=0, padx=5, pady=5)
        fis_l5.grid(row=5, column=0, columnspan=4, padx=5, pady=0)
        fis_l6.grid(row=6, column=0, padx=5, pady=5)
        fis_l7.grid(row=6, column=2, padx=5, pady=5)
        fis_l8.grid(row=7, column=0, padx=5, pady=5)
        fis_l9.grid(row=7, column=2, padx=5, pady=5)
        # --- Entries ---------------------------------------------------------
        self.fis_e1 = Entry(f_fis, width=18)        
        self.fis_e2 = Entry(f_fis, width=18)        
        self.fis_e3 = Entry(f_fis, width=18)        
        self.fis_e4 = Entry(f_fis, width=18)        
        self.fis_e5 = Entry(f_fis, width=18)        
        self.fis_e6 = Entry(f_fis, width=18)        
        self.fis_e7 = Entry(f_fis, width=18)
        self.fis_e1.grid(row=2, column=1)
        self.fis_e2.grid(row=2, column=3)
        self.fis_e3.grid(row=3, column=1)
        self.fis_e4.grid(row=6, column=1)
        self.fis_e5.grid(row=6, column=3)
        self.fis_e6.grid(row=7, column=1)
        self.fis_e7.grid(row=7, column=3)
        # --- Separator -------------------------------------------------------
        fis_s = ttk.Separator(f_fis, orient=HORIZONTAL)
        fis_s.grid(row=4, columnspan=4, sticky=EW, pady=5, padx=5)

        # EXCEL KMZ READY TAB #################################################
        # --- Main Frame ------------------------------------------------------
        tab_kmz = ttk.Frame(n)
        n.add(tab_kmz, text='Excel KMZ')
        # --- inside Frame ----------------------------------------------------
        f_kmz = ttk.Frame(tab_kmz)
        f_kmz.pack_propagate(0)
        f_kmz.pack(padx=25, pady=25)
        # --- Labels ----------------------------------------------------------
        kmz_l1 = Label(f_kmz, text="Icons", bg=bg, fg=fg, font=font)
        kmz_l2 = Label(f_kmz, text="Nº Icon", bg=bg, fg=fg)
        kmz_l3 = Label(f_kmz, text="Escala", bg=bg, fg=fg)
        kmz_l4 = Label(f_kmz, text="Cor", bg=bg, fg=fg)
        kmz_l5 = Label(f_kmz, text=" ", bg=bg, fg=fg)
        kmz_l6 = Label(f_kmz, text="Polígonos", bg=bg, fg=fg, font=font)
        kmz_l7 = Label(f_kmz, text="Raio", bg=bg, fg=fg)
        kmz_l8 = Label(f_kmz, text="Altitude", bg=bg, fg=fg)
        kmz_l9 = Label(f_kmz, text="Cor", bg=bg, fg=fg)
        kmz_l1.grid(row=1, column=0, columnspan=6, padx=5, pady=5)        
        kmz_l2.grid(row=2, column=0, padx=5, pady=5)        
        kmz_l3.grid(row=2, column=2, padx=5, pady=5)        
        kmz_l4.grid(row=2, column=4, padx=5, pady=5)        
        kmz_l5.grid(row=2, column=6, padx=5, pady=5)        
        kmz_l6.grid(row=5, column=0, columnspan=6, padx=5, pady=5)        
        kmz_l7.grid(row=6, column=0, padx=5, pady=5)        
        kmz_l8.grid(row=6, column=2, padx=5, pady=5)        
        kmz_l9.grid(row=6, column=4, padx=5, pady=5)
        # --- Entries ---------------------------------------------------------
        self.kmz_e1 = Entry(f_kmz, width=3)
        self.kmz_e2 = Entry(f_kmz, width=3)
        self.kmz_e3 = Entry(f_kmz)
        self.kmz_e4 = Entry(f_kmz, width=3)
        self.kmz_e5 = Entry(f_kmz, width=3)
        self.kmz_e6 = Entry(f_kmz)
        self.kmz_e1.grid(row=2, column=1)        
        self.kmz_e2.grid(row=2, column=3)        
        self.kmz_e3.grid(row=2, column=5)        
        self.kmz_e4.grid(row=6, column=1)        
        self.kmz_e5.grid(row=6, column=3)        
        self.kmz_e6.grid(row=6, column=5)
        # --- Separator -------------------------------------------------------
        kmz_s = ttk.Separator(f_kmz, orient=HORIZONTAL)
        kmz_s.grid(row=4, columnspan=7, sticky=EW, pady=5, padx=5)

    def __queryDB_button(self):
        all_querys = ''  ######################################################
        msg = 'Pesquisas selecionadas:\n'
        # CGI TAB QUERY -------------------------------------------------------
        cgi_op = self.cgi_e1.get()
        cgi_lac = self.cgi_e2.get()
        cgi_cid = self.cgi_e3.get()
        cgi_cgi = self.cgi_e4.get()
        if cgi_cgi and (cgi_op or cgi_lac or cgi_cid):
            message = "Não é possí­vel pesquisar por CGI em partes e CGI completo ao mesmo tempo."
            messagebox.showerror(title="Erro CGI", message=message)  # falta parent="" e activar a tab CGI
            return          
        else:
            if cgi_op and not cgi_lac and not cgi_cid:
                self.query_dict['cgi'] = "268-" + cgi_op + "%"
                msg += "  - CGI por Operadora\n"
            elif cgi_op and cgi_lac and not cgi_cid:
                self.query_dict['cgi'] = "268-" + cgi_op + "-" + cgi_lac + "%"
                msg += "  - CGI por Operadora e LAC\n"
            elif cgi_op and cgi_cid and not cgi_lac:
                self.query_dict['cgi'] = "268-" + cgi_op + "-" + "%" + "-" + cgi_cid
                msg += "  - CGI por Operadora e CID\n"
            elif cgi_op and cgi_cid and cgi_lac:
                self.query_dict['cgi'] = "268-" + cgi_op + "-" + cgi_lac + "-" + cgi_cid
                msg += "  - CGI por Operadora, LAC e CID\n"
            elif cgi_cid and not cgi_lac and not cgi_op:
                self.query_dict['cgi'] = "%" + "-" + cgi_cid
                msg += "  - CGI por CID\n"
            elif cgi_cid and cgi_lac and not cgi_op:
                self.query_dict['cgi'] = "%" + "-" + cgi_lac + "-" + cgi_cid
                msg += "  - CGI por LAC e CID\n"
            elif cgi_lac and not cgi_op and not cgi_cid:
                self.query_dict['cgi'] = "268-" + "%" + "-" + cgi_lac + "-" + "%"
                msg += "  - CGI por LAC\n"
            elif cgi_cgi:
                self.query_dict['cgi'] = cgi_cgi
                msg += "  - CGI por CGI completo\n"
        # TECNOLOGIA TAB QUERY ------------------------------------------------
        tecs = ''
        tec_lst = [self.tec_2g.get(), self.tec_3g.get(), self.tec_4g.get()]
        for i in tec_lst:
            if i > 0:
                tecs += str(i)
        if tecs:
            self.query_dict['tec'] = tecs
            msg += "  - Por Geração Tecnológica\n"
        # DATA TAB QUERY ------------------------------------------------------
        month_dict = {'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                      'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                      'SET': '09', 'OUT': '10', 'NOV': '11', 'DEC': '12'}
        day_1 = self.dat_cb1.get()
        try:
            month_1 = month_dict[self.dat_cb2.get()]
        except:
            month_1 = None
        year_1 = self.dat_cb3.get()
        day_2 = self.dat_cb4.get()
        try:
            month_2 = month_dict[self.dat_cb5.get()]
        except:
            month_2 = None
        year_2 = self.dat_cb6.get()
        try:
            from_date = day_1 + "/" + month_1 + "/" + year_1
        except:
            from_date = "None"
        try:
            to_date = day_2 + "/" + month_2 + "/" + year_2
        except:
            to_date = "None"
        if (not day_1 and not month_1 and not year_1) and \
           (not day_2 and not month_2 and not year_2):
            pass # Assim está tudo bem, não faz pesquisa por data
        elif (day_1 and month_1 and year_1) and \
             (not day_2 and not month_2 and not year_2):
            pass # Preenche são a data de início
        elif (not day_1 and not month_1 and not year_1) and \
             (day_2 and month_2 and year_2):
            pass # Preenche só a data de fim
        elif (not day_1 or not month_1 or not year_1) and \
            ((not day_2 and not month_2 and not year_2) or \
            (day_2 and month_2 and year_2)):
            message = "A primeira data está incompleta"
            messagebox.showerror(title="Erro Pesquisa por Data", message=message)  # falta parent="" e activar a tab Data
            return
        elif (not day_2 or not month_2 or not year_2):
            message = "A segunda data está incompleta"
            messagebox.showerror(title="Erro Pesquisa por Data", message=message)  # falta parent="" e activar a tab Data
            return
        else:
            if from_date:
                self.query_dict['date']['first'] = from_date
            if to_date:
                self.query_dict['date']['last'] = to_date            
            msg += "  - Por Espaço Temporal\n"
        if self.date_all.get():
                self.query_dict['date']['d_all'] = True
                msg += "  - Incluir todas as actualizações\n"

        # LOCALIZAÇÃO ADMINISTRATIVA TAB QUERY --------------------------------
        pol_cp = self.pol_e1.get()
        pol_loc = self.pol_e2.get()
        pol_con = self.pol_e3.get()
        pol_dis = self.pol_e4.get()
        if pol_cp and not pol_loc and not pol_con and not pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            msg += "  - Loc. Administrativa por Cód Postal\n"
        elif pol_cp and pol_loc and not pol_con and not pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['loc'] = pol_loc
            msg += "  - Loc. Administrativa por Cód Postal e Localidade\n"
        elif pol_cp and not pol_loc and pol_con and not pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['con'] = pol_con
            msg += "  - Loc. Administrativa por Cód Postal e Concelho\n"
        elif pol_cp and not pol_loc and not pol_con and pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Cód Postal e Distrito\n"
        elif pol_cp and pol_loc and pol_con and not pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['con'] = pol_con
            msg += "  - Loc. Administrativa por Cód Postal, Localidade e Concelho\n"
        elif pol_cp and pol_loc and not pol_con and pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Cód Postal, Localidade e Distrito\n"
        elif pol_cp and not pol_loc and pol_con and pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['con'] = pol_con
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Cód. Postal, Concelho e Distrito\n"
        elif pol_cp and pol_loc and pol_con and pol_dis:
            self.query_dict['pol']['cp'] = pol_cp
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['con'] = pol_con
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Cód. Postal, Localidade, Concelho e Distrito\n"
        elif pol_loc and not pol_cp and not pol_con and not pol_dis:
            self.query_dict['pol']['loc'] = pol_loc
            msg += "  - Loc. Administrativa por Localidade\n"
        elif pol_loc and pol_con and not pol_dis and not pol_cp:
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['con'] = pol_con
            msg += "  - Loc. Administrativa por Localidade e Concelho\n"
        elif pol_loc and pol_dis and not pol_cp and not pol_con:
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Localidade e Distrito\n"
        elif pol_loc and pol_con and pol_dis and not pol_cp:
            self.query_dict['pol']['loc'] = pol_loc
            self.query_dict['pol']['con'] = pol_con
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Localidade, Concelho e Distrito\n"
        elif pol_con and not pol_cp and not pol_loc and not pol_dis:
            self.query_dict['pol']['con'] = pol_con
            msg += "  - Loc. Administrativa por Concelho\n"
        elif pol_con and pol_dis and not pol_cp and not pol_loc:
            self.query_dict['pol']['con'] = pol_con
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Concelho e Distrito\n"
        if pol_dis and not pol_cp and not pol_loc and not pol_con:
            self.query_dict['pol']['dis'] = pol_dis
            msg += "  - Loc. Administrativa por Distrito\n"

        # LOCALIZAÇÃO FÍSICA TAB QUERY ----------------------------------------
        fis_lat = self.fis_e1.get()
        fis_lon = self.fis_e2.get()
        fis_rad = self.fis_e3.get()
        fis_lat1 = self.fis_e4.get()
        fis_lon1 = self.fis_e5.get()
        fis_lat2 = self.fis_e6.get()
        fis_lon2 = self.fis_e7.get()
        if (fis_lat or fis_lon or fis_rad) and \
           (fis_lat1 or fis_lon1 or fis_lat2 or fis_lon2):
            message = "Não é possí­vel desenhar círculo e quadrado ao mesmo tempo."
            messagebox.showerror(title="Erro Pesquisa por Localização Física", message=message)  # falta parent="" e activar a tab Loc FÃ­sica
            return
        else:
            if fis_lat and fis_lon and fis_rad:
                self.query_dict['fis']['form'] = 'circle'
                self.query_dict['fis']['lat'] = fis_lat
                self.query_dict['fis']['lon'] = fis_lon
                self.query_dict['fis']['radius'] = fis_rad
                msg += "  - Loc. Física por cí­rculo\n"
            elif fis_lat1 and fis_lon1 and fis_lat2 and fis_lon2:
                self.query_dict['fis']['form'] = 'square'
                self.query_dict['fis']['lat'] = fis_lat1
                self.query_dict['fis']['lon'] = fis_lon1
                self.query_dict['fis']['lat2'] = fis_lat2
                self.query_dict['fis']['lon2'] = fis_lon2
                msg += "  - Loc. Fí­sica por quadrado\n"
        # EXCEL KMZ READY TAB -------------------------------------------------
        kmz_i_n = self.kmz_e1.get()
        kmz_i_sca = self.kmz_e2.get()
        kmz_i_col = self.kmz_e3.get()
        kmz_p_rad = self.kmz_e4.get()
        kmz_p_alt = self.kmz_e5.get()
        kmz_p_col = self.kmz_e6.get()
        if (kmz_i_n or kmz_i_sca or kmz_i_col) and \
           (kmz_p_rad or kmz_p_alt or kmz_p_col):
            message = "Só é possí­vel criar Excel ou de Icones ou de Polí­gonos."
            messagebox.showerror(title="Erro no Excel KMZ", message=message)  # falta parent="" e activar a tab KMZ
            return
        else:
            if kmz_i_n or kmz_i_sca or kmz_i_col:
                self.query_dict['kmz']['ico_or_pol'] = 'icon'
                if kmz_i_n:
                    self.query_dict['kmz']['n'] = kmz_i_n
                if kmz_i_sca:
                    self.query_dict['kmz']['scale'] = kmz_i_sca
                if kmz_i_col:
                    self.query_dict['kmz']['color'] = kmz_i_col
                msg += "  - Excel para KMZ de Icones\n"
            elif kmz_p_rad:
                self.query_dict['kmz']['ico_or_pol'] = 'polygon'
                if kmz_p_alt:
                    self.query_dict['kmz']['alt'] = kmz_p_alt
                if kmz_p_col:
                    self.query_dict['kmz']['color'] = kmz_p_col
                msg += "  - Excel para KMZ de Polígonos\n"


        # DEBUG PRINTS ########################################
        print(msg)
        self.msg.set(msg)
        print()
        print(self.query_dict)


class IconDialog(simpledialog.Dialog):

    def body(self, master):
        # --- Labels ----------------------------------------------------------
        l1 = Label(master, text='ICON')        
        l2 = Label(master, text="Cor:")        
        l3 = Label(master, text="Escala:")        
        l4 = Label(master, text='Retirar CGIs repetidas')        
        l5 = Label(master, text='Tempo (min)')
        l1.grid(row=0, columnspan=2)
        l2.grid(row=1)
        l3.grid(row=2)
        l4.grid(row=3, columnspan=2)
        l5.grid(row=4)
        # --- Entries ---------------------------------------------------------
        self.e1 = Entry(master)
        self.e2 = Entry(master)
        self.e3 = Entry(master)
        self.e1.grid(row=1, column=1)
        self.e2.grid(row=2, column=1)
        self.e3.grid(row=4, column=1)
        #return self.e1 # initial focus

    def apply(self):
        global icon_color, icon_scale, cgi_time
        icon_color = self.e1.get()
        icon_scale = float(self.e2.get())
        cgi_time = int(self.e3.get())


class Controler(object):

    def __init__(self):
        #self.ui = UI()
        #self.ui_paragon = UiParagon()
        #self.ui_filter = UiFilter()
        #self.ui_kmz = UiKmz()
        #self.icon = UiIcon()
        #self.polygon = UiPolygon()
        self.filter_list = []
        self.input_file_name = None

    def db_or_kmz(self):
        choice = 0
        while choice != 2:
            self.ui.db_vs_kmz()
            try:
                choice = int(input(' '))
            except:
                self.db_or_kmz()
            if choice == 1:
                self.ui.db()

    def paragon_data(self):
        choice = ''
        excel_loading_error = True
        paragon = self.ui_paragon
        while not os.path.isfile(choice) or excel_loading_error:
            paragon.open_file()
            choice = input(' ').strip()

            info_paragon = InfoParagon()
            open_xls = info_paragon.open_xls(choice)

            if open_xls != 'Error':
                excel_loading_error = False
            else:
                paragon.wrong_file()

        self.input_file_name = choice
        paragon.correct_file()

    def __filter_cgi(self):
        cgi = ''
        while cgi != 's' and cgi != 'n':
            self.ui_filter.repeated_cgis_time()
            cgi = input(' ').lower().strip()

        mins = ''
        if cgi == 's':
            while not mins.isnumeric():
                self.ui_filter.time_to_remove_cgis()
                mins = input(' ').lower().strip()
        else:
            mins = 0

        return int(mins)

    def __filter_operator(self):
        '''
        ope = "e"  # qq coisa p não ser uma string vazia
        OPE = "mvn"
        ope_input_error = True

        while ope_input_error:
            for char in ope:
                if char in OPE:
                    ope_input_error = False
                else:
                    ope_input_error = True
                    self.ui_filter.operator()
                    ope = input(' ').lower().strip()
                    break

        return ope'''
        return "mvn"

    def __filter_technology(self):
        '''
        tec = 'e'
        TEC = '234'
        tec_input_error = True

        while tec_input_error or not tec.isnumeric():
            for char in tec:
                if char in TEC:
                    tec_input_error = False
                else:
                    tec_input_error = True
                    self.ui_filter.technology()
                    tec = input(' ').strip()
                    break

        return tec'''
        return '234'

    '''def __filter_neighbor_cells(self):
        neighbor_cells = ''
        while neighbor_cells != 's' and neighbor_cells != 'n':
            self.ui_filter.neighbor_cells()
            neighbor_cells = input(' ').lower().strip()

        radius = ''
        if neighbor_cells == 's':
            while not radius.isnumeric() or radius != '5' and \
                    radius != '20' and radius != '40':
                self.ui_filter.neighbor_cells_radius()
                radius = input(' ').strip()
        else:
            radius = '0'

        color = None
        if neighbor_cells == 's':
            while color not in Colors().get_colors():
                self.ui_filter.neighbor_cells_color()
                color = input(' ').lower().strip()

        return int(radius), color'''

    def __icon_number(self):
        '''
        first verify if the input is a number, than if that number isn't
        between 1 and 579 call the same function again.
        '''
        '''number = ''

        while not number.isnumeric():
            self.icon.number()
            number = input(' ').strip()

        if int(number) < 1 or int(number) > 579:
            self.__icon_number()

        return int(number)'''
        return 338

    def __icon_color(self, color):
        '''MODIFIED FOR SIMPLE GUI!!!'''
        '''
        color = ''
        while not color in Colors().get_colors():
            self.icon.color()
            color = input(' ').lower().strip()
        '''
        return color

    def __icon_scale(self, scale):
        '''MODIFIED FOR SIMPLE GUI!!!'''
        '''
        scale = ''
        try:
            float(scale)
        except ValueError:
            self.icon.scale()
            scale = input(' ').lower().strip()

        if float(scale) < 0.1:
            self.__icon_scale()'''

        return float(scale)

    def __icon_line_string(self):
        color = None
        '''
        use_line = ''
        while use_line != 's' and use_line != 'n':
            self.icon.use_line()
            use_line = input(' ').lower().strip()

        if use_line == 's':
            while not color in Colors().get_colors():
                self.icon.color()
                color = input(' ').lower().strip()'''

        return color

    def __polygon_amplitude(self):
        amplitude = ''
        while not amplitude.isnumeric():
            self.polygon.amplitude()
            amplitude = input(' ').strip()

        if int(amplitude) < 1 or int(amplitude) > 360:
            self.__polygon_amplitude()

        return int(amplitude)

    def __polygon_radius(self):
        radius = ''
        while not radius.isnumeric():
            self.polygon.radius()
            radius = input(' ').strip()

        if int(radius) < 1:
            self.__polygon_radius()

        return int(radius)

    def __polygon_altitude(self):
        altitude = ''
        while not altitude.isnumeric():
            self.polygon.altitude()
            altitude = input(' ').strip()

        if int(altitude) < 1:
            self.__polygon_altitude()

        return altitude

    def __polygon_color(self):
        color = ''
        while color not in Colors().get_colors():
            self.polygon.color()
            color = input(' ').strip()

        return color

    def icon_or_polygon(self, color, scale):
        '''
        MODIFIED FOR THE SIMPLE GUI!!!!!!!!!
        kmz = ''
        while kmz != 'i' and kmz != 'p':
            UiKmz().icon_vs_polygon()
            kmz = input(' ').lower().strip()
        '''
        kmz = 'i'

        if kmz == 'i':
            number = self.__icon_number()
            color = self.__icon_color(color)
            scale = self.__icon_scale(scale)
            line_color = self.__icon_line_string()

            icon = Icon()
            icon.set_n_icon(number)
            icon.set_color(color)
            icon.set_scale(scale)
            icon.set_line_string(line_color)
            icon.set_file_name = self.input_file_name
            icon.set_filter_list = self.filter_list

            return icon

        else:
            amplitude = self.__polygon_amplitude()
            radius = self.__polygon_radius()
            altitude = self.__polygon_altitude()
            color = self.__polygon_color()

            polygon = Polygon()
            polygon.set_amplitude(amplitude)
            polygon.set_radius(radius)
            polygon.set_altitude(altitude)
            polygon.set_color(color)
            polygon.set_file_name = self.input_file_name
            polygon.set_filter_list = self.filter_list

            return polygon

    def filter_choices(self):
        _filter = self.ui_filter

        filter_cgi = self.__filter_cgi()
        self.filter_list.append(filter_cgi)
        _filter.cgis_time_applied()

        filter_ope = self.__filter_operator()
        self.filter_list.append(filter_ope)
        #_filter.operator_applied()

        filter_tec = self.__filter_technology()
        self.filter_list.append(filter_tec)
        #_filter.technology_applied()

        #filter_neighbor_cells = self.__filter_neighbor_cells()
        #self.filter_list.append(filter_neighbor_cells[0])  # kms
        #self.filter_list.append(filter_neighbor_cells[1])  # cor, None se Kms=0
        self.filter_list.append('n')  # delete in the future
        self.filter_list.append(None)  # delete in the future
        #_filter.neighbor_cells_applied()

        return self.filter_list


#==============================================================================
# CLASSES FROM BUSINESS PACKAGE
#==============================================================================
class CGI(object):

    def __init__(self):
        self.cgi = None
        self.latitude = None
        self.longitude = None
        self.address = None
        self.local = None
        self.name = None
        self.zip = None
        self.azimute = None
        self.technology = None
        self.date = None

    def set_cgi(self, cgi):
        self.cgi = cgi

    def set_lat(self, lat):
        self.latitude = lat

    def set_lon(self, lon):
        self.longitude = lon

    def set_address(self, address):
        self.address = address

    def set_local(self, local):
        self.local = local

    def set_name(self, name):
        self.name = name

    def set_zip(self, _zip):
        self.zip = _zip

    def set_azimute(self, azimute):
        self.azimute = azimute

    def set_technology(self, technology):
        self.technology = technology

    def set_date(self, _date):
        self.date = _date

    def get_cgi(self):
        return self.cgi

    def get_lat(self):
        return self.latitude

    def get_lon(self):
        return self.longitude

    def get_address(self):
        return self.address

    def get_local(self):
        return self.local

    def get_name(self):
        return self.name

    def get_zip(self):
        return self.zip

    def get_azimute(self):
        return self.azimute

    def get_technology(self):
        return self.technology

    def get_date(self):
        return self.date


class InfoCgi(object):

    def __init__(self):
        self.cgi = CGI()
        self.query = None
        self.result = None

    def get_query_db(self, query):
        db_query = SingleQuery(query)
        result = db_query.cgi()[0]
        if result:
            self.result = result[0]
            # to get the result with most recent date
            if len(result) > 1:
                self.__get_most_recent_date(result)
        else:
            self.result = None

    def __get_most_recent_date(self, result):
        result_copy = result[:]

        for item in result:
            if item[9] > self.result[9]:
                self.result = result_copy[0]
            result_copy.pop(0)

    def build_cgi(self):
        '''
        self.result is a tuple of query results. Ex:
        ('268-06-8000-31756', 38.67376, -9.176895, 'AVENIDA TORRADO SILVA',
        'PRAGAL', 'HOSPITAL GARCIA ORTA FDD 2', '2805-267', 5, '3G',
        '2014-05-26')
        '''
        if self.result:
            self.cgi.set_cgi(self.result[0])
            self.cgi.set_lat(self.result[1])
            self.cgi.set_lon(self.result[2])
            self.cgi.set_address(self.result[3])
            self.cgi.set_local(self.result[4])
            self.cgi.set_name(self.result[5])
            self.cgi.set_zip(self.result[6])
            self.cgi.set_azimute(self.result[7])
            self.cgi.set_technology(self.result[8])
            self.cgi.set_date(self.result[9])
        return self.cgi


class Colors(object):

    def __init__(self):
        self.colors = ['aliceblue', 'antiquewhite', 'aqua', 'aquamarine',
                       'azure', 'beige', 'bisque', 'black', 'blanchedalmond',
                       'blue', 'blueviolet', 'brown', 'burlywood', 'cadetblue',
                       'chartreuse', 'chocolate', 'coral', 'cornflowerblue',
                       'cornsilk', 'crimson', 'cyan', 'darkblue', 'darkcyan',
                       'darkgoldenrod', 'darkgreen', 'darkgrey', 'darkkhaki',
                       'darkmagenta', 'darkolivegreen', 'darkorange',
                       'darkorchid', 'darkred', 'darksalmon', 'darkseagreen',
                       'darkslateblue', 'darkslategray', 'darkturquoise',
                       'darkviolet', 'deeppink', 'deepskyblue', 'dimgray',
                       'dodgerblue', 'firebrick', 'floralwhite', 'forestgreen',
                       'fuchsia', 'gainsboro', 'ghostwhite', 'gold',
                       'goldenrod', 'gray', 'green', 'greenyellow', 'honeydew',
                       'hotpink', 'indianred', 'indigo', 'ivory', 'khaki',
                       'lavender', 'lavenderblush', 'lawngreen',
                       'lemonchiffon', 'lightblue', 'lightcoral', 'lightcyan',
                       'lightgoldenrodyellow', 'lightgray', 'lightgreen',
                       'lightpink', 'lightsalmon', 'lightseagreen',
                       'lightskyblue', 'lightslategray', 'lightsteelblue',
                       'lightyellow', 'lime', 'limegreen', 'linen', 'magenta',
                       'maroon', 'mediumaquamarine', 'mediumblue',
                       'mediumorchid', 'mediumpurple', 'mediumseagreen',
                       'mediumslateblue', 'mediumspringgreen',
                       'mediumturquoise', 'mediumvioletred', 'midnightblue',
                       'mintcream', 'mistyrose', 'moccasin', 'navajowhite',
                       'navy', 'oldlace', 'olive', 'olivedrab', 'orange',
                       'orangered', 'orchid', 'palegoldenrod', 'palegreen',
                       'paleturquoise', 'palevioletred', 'papayawhip',
                       'peachpuff', 'peru', 'pink', 'plum', 'powderblue',
                       'purple', 'red', 'rosybrown', 'royalblue',
                       'saddlebrown', 'salmon', 'sandybrown', 'seagreen',
                       'seashell', 'sienna', 'silver', 'skyblue', 'slateblue',
                       'slategray', 'snow', 'springgreen', 'steelblue', 'tan',
                       'teal', 'thistle', 'tomato', 'turquoise', 'violet',
                       'wheat', 'white', 'whitesmoke', 'yellow', 'yellowgreen']

    def get_colors(self):
        return self.colors


class BusinessLogic(object):
    
    #==========================================================================
    # CLASS CONSTRUCTOR
    #==========================================================================

    def __init__(self):
        # Excel I/O related variables -----------------------------------------
        self.xls = InfoParagon()
        self.final_excel = XlsWriter()
        self.no_results_excel = XlsWriter()

        # Filter related variables --------------------------------------------
        self.filter_list = []
        self.first_date = None

        # Paragon related variables -------------------------------------------
        self.line = []  # raw paragon line
        self.paragon_titles_row = []
        # [tipo, direcção, nº produto, hora, dia, cgi]
        self.paragon_line_1 = []  # call start
        self.paragon_line_2 = []  # call end
        #self.paragon_previous_line = []
        self.previous_line_1 = []
        self.previous_line_2 = []
        self.last_cgi = None
        self.filter_cgi = False

        # DB related variables -----------------------------------------------
        self.cgi_db_line_1 = None
        self.cgi_db_line_2 = None

        # No results related variables ---------------------------------------
        self.use_no_results_excel = False
        self.no_results_counter = 0
        self.prev_final_line = []
        self.post_final_lines = []
        self.append_to_post = False

    #==========================================================================
    # PUBLIC FUNCTIONS / METHODS
    #==========================================================================

    def build_icon_file(self, icon, filter_list):
        '''
        icon is an Icon() object
        filter_list is a list with:
            [0] - cgis_to_remove_by_time_in_minutes - numeric str
            [1] - operator - str (m/v/o)
            [2] - technology - numeric str
            [3] - neighbour_cells - str (s/n)
            [4] - neighbour_cells_color - str (color) / None
        '''
        self.filter_list = filter_list
        self.xls.open_xls(icon.get_file_name())
        worksheet_names = self.xls.get_worksheet_names()
        self.xls.workbook = self.xls.get_workbook()

        for sheet_name in worksheet_names:
            self.__prepare_excel_final_lines(icon, sheet_name)

        file_name = self.final_excel.save_workbook(icon.get_file_name(), 'icon')
        # PRECISO DE SABER O NOME DESTE FICHEIRO NOVO PARA ENFIAR NO CGITIMEFILTER()
        if self.use_no_results_excel:
            self.__write_post_lines()
            self.no_results_excel.save_workbook(icon.get_file_name(), 'no_res')
        #if self.filter_cgi:  # commented 13FEV15
        CgiTimeFilter(file_name, int(self.filter_list[0])).filter()

    def build_polygon_file(self, polygon, filter_list):
        '''
        polygon is an Polygon() object
        filter_list is a list with:
            [0] - cgis_to_remove_by_time_in_minutes - numeric str
            [1] - operator - str (m/v/o)
            [2] - technology - numeric str
            [3] - neighbour_cells - str (s/n)
            [4] - neighbour_cells_color - str (color) / None
        '''
        self.filter_list = filter_list
        self.xls.open_xls(polygon.get_file_name())
        worksheet_names = self.xls.get_worksheet_names()
        self.xls.workbook = self.xls.get_workbook()

        for sheet_name in worksheet_names:
            self.__prepare_excel_final_lines(polygon, sheet_name)

        file_name = self.final_excel.save_workbook(polygon.get_file_name(), 
                                                   'polygon')
        if self.use_no_results_excel:
            self.no_results_excel.save_workbook(polygon.get_file_name(),
                                                'no_res')
        if self.filter_cgi:
            CgiTimeFilter(file_name, int(self.filter_list[0])).filter()

    #==========================================================================
    # PRIVATE / HELPER FUNCTIONS / METHODS -  EXCEL BUILD RELATED
    #==========================================================================

    def __prepare_excel_final_lines(self, icon_or_polygon, sheet_name):
        '''
        icon_or_polygon - is an Icon() or Polygon() object
        '''
        date_index = None
        # Build First ExcelFinal line -----------------------------------------
        if isinstance(icon_or_polygon, Icon):
            if icon_or_polygon.get_line_string() == None:
                fst_line = self.__icon_first_line()
            else:
                fst_line = self.__icon_first_line_with_linestring()
        else:
            fst_line = self.__polygon_first_line()

        # Prepare Paragon Excel -----------------------------------------------
        self.xls.worksheet = self.xls.workbook.sheet_by_name(sheet_name)
        self.xls.open_worksheet(sheet_name)
        num_rows = self.xls.get_n_rows()  # total number of rows
        self.xls.num_cells = self.xls.worksheet.ncols - 1  # cols total

        # Manipulate the Excel Lines ------------------------------------------
        n_sheets = 0
        for line in range(num_rows + 1):
            self.line = self.xls.get_next_line()  # raw paragon line

            lower_line = [str(x).lower() for x in self.line]

            if lower_line[0] == 'tipo de produto':
                self.paragon_titles_row = lower_line.copy()
                date_index = self.paragon_titles_row.index('hora inicial')
                self.__build_sheet_name(date_index)
                self.final_excel.write_line(fst_line)

            if self.line[date_index][:10] != self.first_date:  # new sheet
                if self.first_date < self.line[date_index][:10]:       
                    self.first_date = self.line[date_index][:10] 
                self.__build_sheet_name(date_index)
                self.final_excel.write_line(fst_line)
                # adicionado 4 jul 14 para o programa meter sempre a primeira
                # linha de cada dia -------------------------------------------
                self.previous_line_1 = []
                self.previous_line_2 = []
                # -------------------------------------------------------------
                self.xls.reset_counter()  # reset line counter            
                n_sheets += 1
            else:
                self.__manipulate_normal_line(icon_or_polygon)
        if n_sheets > 0:
            self.__excel_final_end_line(icon_or_polygon, n_sheets, date_index,
                                        fst_line)

    def __excel_final_end_line(self, icon_or_polygon, n_sheets, date_index,
                               fst_line):
        new_n_sheets = 0
        for line in range(n_sheets):
            self.line = self.xls.get_next_line()  # raw paragon line

            lower_line = [str(x).lower() for x in self.line]

            if lower_line[0] == 'tipo de produto':
                self.paragon_titles_row = lower_line.copy()
                date_index = self.paragon_titles_row.index('hora inicial')
                self.__build_sheet_name(date_index)
                self.final_excel.write_line(fst_line)

            elif self.line[date_index][:10] != self.first_date:  # new sheet
                if self.first_date < self.line[date_index][:10]:       
                    self.first_date = self.line[date_index][:10]
                self.__build_sheet_name(date_index)
                self.final_excel.write_line(fst_line)
                # adicionado 4 jul 14 para o programa meter sempre a primeira
                # linha de cada dia -------------------------------------------
                self.previous_line_1 = []
                self.previous_line_2 = []
                # -------------------------------------------------------------
                self.xls.reset_counter()  # reset line counter
                new_n_sheets += 1
            else:
                self.__manipulate_normal_line(icon_or_polygon)

        if new_n_sheets > 0:
            self.__excel_final_end_line(icon_or_polygon, new_n_sheets,
                                        date_index, fst_line)

    def __manipulate_normal_line(self, icon_or_polygon):
        '''
        icon_or_polygon - is an Icon() or Polygon() object
        '''
        # build two paragon lines from the raw line ---------------------------
        self.__paragon_line()

        # apply filters (cgi_time, operator, technology, neighbor)
        self.__apply_paragon_filters()

        # query db to construct db lines --------------------------------------
        if self.__not_empty(self.paragon_line_1):
            cgi = self.paragon_line_1[5]
            self.cgi_db_line_1 = self.__cgi_from_db(cgi)
        if self.__not_empty(self.paragon_line_2):
            cgi = self.paragon_line_2[5]
            self.cgi_db_line_2 = self.__cgi_from_db(cgi)

        # filter db lines -----------------------------------------------------
        if self.cgi_db_line_1 or self.cgi_db_line_2:
            self.__apply_db_filters()

        # Build an ExcelFinal line and write it with ExcelWriter or
        # Build no results file in case DB returns Null -----------------------
        self.__build_excel_final_lines(icon_or_polygon)

        # Update self.paragon_previous_line -----------------------------------
        # Alterado 26 JUN 14
        if self.__not_empty(self.paragon_line_2) and \
                len(self.paragon_line_2[-1]) > 0:
            self.previous_line_2 = self.paragon_line_2.copy()
        if self.__not_empty(self.paragon_line_1) and \
                len(self.paragon_line_1[-1]) > 0:
            self.previous_line_1 = self.paragon_line_1.copy()

    def __build_excel_final_lines(self, icon_or_polygon):
        '''
        icon_or_polygon - is an Icon() or Polygon() object
        '''
        # paragon line 1 ------------------------------------------------------
        if self.__not_empty(self.paragon_line_1):
            if self.cgi_db_line_1.get_lat():
                paragon_1 = self.paragon_line_1.copy()
                db_cgi_1 = self.cgi_db_line_1
                final_1 = ExcelFinal(icon_or_polygon, paragon_1, db_cgi_1)
                final_1.set_line()
                line_1 = final_1.get_line()
                if self.append_to_post:
                    final_lines = len(self.post_final_lines)
                    for i in range(self.no_results_counter - final_lines):
                        self.post_final_lines.append(line_1)
                    self.append_to_post = False
                self.final_excel.write_line(line_1)
                self.prev_final_line = line_1.copy()
                self.last_cgi = line_1[3]
            else:
                self.__no_results_file(self.paragon_line_1, icon_or_polygon)

        # paragon line 2 ------------------------------------------------------
        if self.__not_empty(self.paragon_line_2):
            if self.cgi_db_line_2.get_lat():
                paragon_2 = self.paragon_line_2.copy()
                db_cgi_2 = self.cgi_db_line_2
                final_2 = ExcelFinal(icon_or_polygon, paragon_2, db_cgi_2)
                final_2.set_line()
                line_2 = final_2.get_line()
                if self.append_to_post:
                    final_lines = len(self.post_final_lines)
                    for i in range(self.no_results_counter - final_lines):
                        self.post_final_lines.append(line_2)
                    self.append_to_post = False
                self.final_excel.write_line(line_2)
                self.prev_final_line = line_2.copy()
                self.last_cgi = line_2[3]
            else:
                self.__no_results_file(self.paragon_line_2, icon_or_polygon)

    def __no_results_file(self, no_results_line, icon_or_polygon):

        if not self.use_no_results_excel:
            fst_line = self.__no_results_first_line()
            self.no_results_excel.add_sheet(date.today().isoformat())
            self.no_results_excel.write_line(fst_line)
            self.use_no_results_excel = True
            self.__no_results_file(no_results_line, icon_or_polygon)
        else:
            self.no_results_counter += 1
            excel = NoResults()
            excel.set_line(no_results_line, self.prev_final_line)
            self.no_results_excel.write_line(excel.get_line())
            self.append_to_post = True

    def __write_post_lines(self):
        self.no_results_excel.row_counter = 2
        for value in self.post_final_lines:
            excel = PostNoResults()
            excel.set_line(value)
            self.no_results_excel.write_line(excel.get_line(), col=13)

    def __build_sheet_name(self, date_index):        
        if self.line[0] == 'Tipo de produto':
            self.line = self.xls.get_next_line()
            self.first_date = self.line[date_index][:10]
        else:
            try:
                self.line = self.xls.get_next_line()
            except:
                pass
        sheet_name = self.__get_week_days(self.first_date)
        self.final_excel.add_sheet(sheet_name)
        self.xls.reset_counter()  # reset line counter

    def __get_week_days(self, _date):
        '''
        date.weekday()
        Return the day of the week as an integer, where Monday is 0
        and Sunday is 6.
        For example, date(2002, 12, 4).weekday() == 2, a Wednesday
        '''
        week_day = ['2_', '3_', '4_', '5_', '6_', 'S_', 'D_']
        day, month, year = (x for x in _date[:10].split('.'))
        ext_month = {'01': 'JAN', '02': 'FEB', '03': 'MAR', '04': 'ABR',
                     '05': 'MAI', '06': 'JUN', '07': 'JUL', '08': 'AGO',
                     '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEC'}

        f_day = week_day[date(int(year), int(month), int(day)).weekday()]

        return f_day + day + ext_month[month]

    def __paragon_line(self):
        '''
        Builds two paragon_lines with:
        [tipo, direcçãoo, nº produto, hora, dia, cell cgi]
        '''
        tipo_i = self.paragon_titles_row.index('tipo de produto')
        dir_i = self.paragon_titles_row.index('direcção')
        prod_i = self.paragon_titles_row.index('produto nº')
        gdh_1_i = self.paragon_titles_row.index('hora inicial')
        gdh_2_i = self.paragon_titles_row.index('hora final')
        cgi_1_i = self.paragon_titles_row.index('cell inicial')
        cgi_2_i = self.paragon_titles_row.index('cell final')

        # empty the paragon_lines
        self.paragon_line_1 = []
        self.paragon_line_2 = []
        # tipo produto [0]         ---> balão_kmz[7] Tipo
        self.paragon_line_1.append(self.line[tipo_i])
        self.paragon_line_2.append(self.line[tipo_i])
        # direcção [1]             ---> balão_kmz[8] Direcção
        self.paragon_line_1.append(self.line[dir_i])
        self.paragon_line_2.append(self.line[dir_i])
        # produto nº [3]           ---> balão_kmz[9] Produto
        self.paragon_line_1.append(self.line[prod_i])
        self.paragon_line_2.append(self.line[prod_i])
        # hora [4][11:] e [5][11:] ---> balão_kmz[0] Name - tb NAME do Ponto
        self.paragon_line_1.append(self.line[gdh_1_i][11:])
        self.paragon_line_2.append(self.line[gdh_2_i][11:])
        # dia [4][:10] e [5][:10]  ---> balão_kmz[1] Dia
        self.paragon_line_1.append(self.line[gdh_1_i][:10])
        self.paragon_line_2.append(self.line[gdh_2_i][:10])
        # cell [10] e [11]         ---> DESCRIPTION do ponto
        self.paragon_line_1.append(self.line[cgi_1_i])
        self.paragon_line_2.append(self.line[cgi_2_i])


    def __not_empty(self, _list): # não preciso disto basta fazer if _list se tiver alguma coisa retorna True, vazia False
        '''
        check if a list is or not empty
        '''
        if len(_list) > 0:
            return True
        return False

    def __cgi_from_db(self, cgi):
        '''
        cria lista com os campos do excel do paragon
        '''
        query = InfoCgi()
        query.get_query_db([cgi])

        return query.build_cgi()

    def __icon_first_line_with_linestring(self):
        return ['Latitude', 'Longitude', 'Name', 'Description',
                'AppendDataColumnsToDescription', 'Icon', 'Iconcolor',
                'IconScale', 'IconHeading', 'LineStringColor', 'Data',
                'Morada', 'Local', 'Nome', 'Azimute', 'Tecnologia', 'Tipo',
                'Direcção', 'Produto']

    def __icon_first_line(self):
        return ['Latitude', 'Longitude', 'Name', 'Description',
                'AppendDataColumnsToDescription', 'Icon', 'Iconcolor',
                'IconScale', 'IconHeading', 'Data', 'Morada', 'Local', 'Nome',
                'Azimute', 'Tecnologia', 'Tipo', 'Direcção', 'Produto']

    def __polygon_first_line(self):
        return ['Latitude', 'Longitude', 'Name', 'Description',
                'AppendDataColumnsToDescription', 'Polygon', 'PolygonColor',
                'PolygonAzimute', 'PolygonAltitude', 'PolygonAmplitude',
                'Data', 'Morada', 'Local', 'Nome', 'Azimute', 'Tecnologia',
                'Tipo', 'Direcção', 'Produto']

    def __no_results_first_line(self):
        return['CGI', 'Data', 'Hora', 'Tipo', 'DireÃ§Ã£o', 'Produto',
               'CGI anterior', 'Data', 'Hora', 'Nome', 'Azimute',
               'Coordenadas', 'CGI Posterior', 'Data', 'Hora', 'Nome',
               'Azimute', 'Coordenadas']

    #==========================================================================
    # PRIVATE / HELPER FUNCTIONS / METHODS - FILTER PARAGON RELATED
    #==========================================================================

    def __apply_paragon_filters(self):
        '''
        self.filter_list = ['cgi_time', 'operator', 'technology',
                            'neighbour_cells', 'n_cells_color']
        '''
        if int(self.filter_list[0]) != 0:
            self.filter_cgi = True
        if len(self.filter_list[1]) != 3:
            self.__filter_operator()

    def __filter_operator(self):
        '''
        self.filter_list = ['cgi_time', 'operator', 'technology',
                            'neighbour_cells', 'n_cells_color']
        MEO ------ 268-06 ---> cgi[:6]
        Vodafone - 268-01 ---> cgi[:6]
        Vodafone - 26801  ---> cgi[:5]
        Optimus -- 268-03 ---> cgi[:6]
        '''
        # paragon_line_1 ------------------------------------------------------
        if self.__not_empty(self.paragon_line_1):
            cgi = self.paragon_line_1[5]
            # use two operators -----------------------------------------------
            if len(self.filter_list[1]) == 2:
                if 'm' in self.filter_list[1] and 'v' in self.filter_list[1]:
                    # ignorar cgi começados por 268-03 ------------------------
                    if self.__operator_is_nos(cgi):
                        self.paragon_line_1 = []
                elif 'm' in self.filter_list[1] and 'n' in self.filter_list[1]:
                    # ignorar cgi começados por 268-01 e 26801 ----------------
                    if self.__operator_is_vodafone(cgi):
                        self.paragon_line_1 = []
                else:
                    # ignorar cgi começados por 268-06 ------------------------
                    if self.__operator_is_meo(cgi):
                        self.paragon_line_1 = []
            # use a single operator -------------------------------------------
            else:
                if 'm' in self.filter_list[1]:
                    # só meter cgi começados por 268-06 -----------------------
                    if not self.__operator_is_meo(cgi):
                        self.paragon_line_1 = []
                elif 'v' in self.filter_list[1]:
                    # só meter cgi começados por 268-01 ou 26801 --------------
                    if not self.__operator_is_vodafone(cgi):
                        self.paragon_line_1 = []
                else:
                    # só meter cgi começados por 268-03 -----------------------
                    if not self.__operator_is_nos(cgi):
                        self.paragon_line_1 = []
        # paragon_line_2 ------------------------------------------------------
        if self.__not_empty(self.paragon_line_2):
            cgi = self.paragon_line_2[5]
            # use two operators ----------------------------------------------
            if len(self.filter_list[1]) == 2:
                if 'm' in self.filter_list[1] and 'v' in self.filter_list[1]:
                    # ignorar cgi começados por 268-03 ------------------------
                    if self.__operator_is_nos(cgi):
                        self.paragon_line_2 = []
                elif 'm' in self.filter_list[1] and 'n' in self.filter_list[1]:
                    # ignorar cgi começados por 268-01 e 26801 ----------------
                    if self.__operator_is_vodafone(cgi):
                        self.paragon_line_2 = []
                else:
                    # ignorar cgi começados por 268-06 ------------------------
                    if self.__operator_is_meo(cgi):
                        self.paragon_line_2 = []
            # use a single operator -------------------------------------------
            else:
                if 'm' in self.filter_list[1]:
                    # só meter cgi começados por 268-06 -----------------------
                    if not self.__operator_is_meo(cgi):
                        self.paragon_line_2 = []
                elif 'v' in self.filter_list[1]:
                    # só meter cgi começados por 268-01 ou 26801 --------------
                    if not self.__operator_is_vodafone(cgi):
                        self.paragon_line_2 = []
                else:
                    # só meter cgi começados por 268-03 -----------------------
                    if not self.__operator_is_nos(cgi):
                        self.paragon_line_2 = []

    def __operator_is_meo(self, cgi):
        '''
        check if the operator is MEO
        '''
        return '268-06' in cgi[:6]

    def __operator_is_vodafone(self, cgi):
        '''
        check if the operator is Vodafone
        '''
        rem = '-'
        cgi_2 = cgi.translate(str.maketrans(dict.fromkeys(rem)))

        return '26801' in cgi_2[:5]

    def __operator_is_nos(self, cgi):
        '''
        check if the operator is NOS
        '''
        return '268-03' in cgi[:6]

    #==========================================================================
    # PRIVATE / HELPER FUNCTIONS / METHODS - FILTER DB RELATED
    #==========================================================================

    def __apply_db_filters(self):
        '''
        self.filter_list = ['cgi_time', 'operator', 'technology',
                            'neighbour_cells', 'n_cells_color']
        '''
        if len(self.filter_list[2]) != 3:
            self.__filter_technology()
        if self.filter_list[3] == 's':
            self.__neighbour_cells()

    def __filter_technology(self):
        '''
        self.filter_list = ['cgi_time', 'operator', 'technology',
                            'neighbour_cells', 'n_cells_color']
        2G = GSM
        3G = HSDPA, UMTS
        4G = FDD_1, LTE, LTE_1800, LTE_2600, LTE_800
        atenção vodafone não tem tecnologia
        '''

        _2g = ['2G', 'GSM']
        _3g = ['3G', 'HSDPA', 'UMTS']
        _4g = ['4G', 'FDD_1', 'LTE', 'LTE_1800', 'LTE_2600', 'LTE_800']

        # cgi_db_line_1 -------------------------------------------------------
        if self.cgi_db_line_1:
            tec = self.cgi_db_line_1.get_technology()
            # use two technologies --------------------------------------------
            if len(self.filter_list[2]) == 2:
                if '2' in self.filter_list[2] and '3' in self.filter_list[2]:
                    # ignore 4G, FDD_1, LTE, LTE_1800, LTE_2600, LTE_800 ------
                    if tec in _4g:
                        self.cgi_db_line_1 = None
                elif '3' in self.filter_list[2] and '4' in self.filter_list[2]:
                    # ignore 2G, GSM ------------------------------------------
                    if tec in _2g:
                        self.cgi_db_line_1 = None
                else:
                    # ignore 3G, HSDPA, UMTS ----------------------------------
                    if tec in _3g:
                        self.cgi_db_line_1 = None
            else:
                # use one technology ------------------------------------------
                if '2' in self.filter_list[2]:
                    # only add line if it has 2G or GSM -----------------------
                    if tec not in _2g:
                        self.cgi_db_line_1 = None
                elif '3' in self.filter_list[2]:
                    # only add line if it has 3G, HSDPA, UMTS -----------------
                    if tec not in _3g:
                        self.cgi_db_line_1 = None
                else:
                    # only add line if it has 4G, FDD_1, LTE, LTE_1800
                    # LTE_2600, LTE_800 ---------------------------------------
                    if tec not in _4g:
                        self.cgi_db_line_1 = None

        # cgi_db_line_2 -------------------------------------------------------
        if self.cgi_db_line_2:
            tec = self.cgi_db_line_2.get_technology()
            # use two technologies --------------------------------------------
            if len(self.filter_list[2]) == 2:
                if '2' in self.filter_list[2] and '3' in self.filter_list[2]:
                    # ignore 4G, FDD_1, LTE, LTE_1800, LTE_2600, LTE_800 ------
                    if tec in _4g:
                        self.cgi_db_line_2 = None
                elif '3' in self.filter_list[2] and '4' in self.filter_list[2]:
                    # ignore 2G, GSM ------------------------------------------
                    if tec in _2g:
                        self.cgi_db_line_2 = None
                else:
                    # ignore 3G, HSDPA, UMTS ----------------------------------
                    if tec in _3g:
                        self.cgi_db_line_2 = None
            else:
                # use one technology ------------------------------------------
                if '2' in self.filter_list[2]:
                    # only add line if it has 2G or GSM -----------------------
                    if tec not in _2g:
                        self.cgi_db_line_2 = None
                elif '3' in self.filter_list[2]:
                    # only add line if it has 3G, HSDPA, UMTS -----------------
                    if tec not in _3g:
                        self.cgi_db_line_2 = None
                else:
                    # only add line if it has 4G, FDD_1, LTE, LTE_1800
                    # LTE_2600, LTE_800 ---------------------------------------
                    if tec not in _4g:
                        self.cgi_db_line_2 = None

    def __neighbour_cells(self):
        '''
        for future implementation
        '''
        raise NotImplementedError


class ExcelFinal(object):

    def __init__(self, icon_or_polygon, paragon_line, db_cgi_query):
        self.line = ExcelLine()

        # (n_icon, scale, heading, line_string, color, file_name)
        self.icon = None

        # (amplitude, altitude, radius, color, file_name)
        self.polygon = None

        # [tipo, direcção, nº produto, hora, dia, cgi]
        self.paragon_line = paragon_line

        # (cgi, latitude, longitude, address, local, name, zip, azimuth,
        # technology, date)
        self.db_line = db_cgi_query

        if isinstance(icon_or_polygon, Icon):
            self.icon = icon_or_polygon
        else:
            self.polygon = icon_or_polygon

        # azimuth to help building the icon heading or polygon azimuth
        self.azimute = None

    def set_line(self):
        '''
        '''
        lat = self.db_line.get_lat()
        lon = self.db_line.get_lon()
        name = self.paragon_line[3]
        description = self.db_line.get_cgi()
        adctd = 'Name, Data, Morada, Nome, Azimute, Tecnologia, Tipo, Direcção, Produto, Coordenadas'
        data = self.paragon_line[4]
        morada = self.db_line.get_address()
        local = self.db_line.get_local()
        nome = self.db_line.get_name()
        self.azimute = self.db_line.get_azimute()
        tec = self.db_line.get_technology()
        tipo = self.paragon_line[0]
        direcao = self.paragon_line[1]
        produto = self.paragon_line[2]

        self.line.set_value(lat)
        self.line.set_value(lon)
        self.line.set_value(name)
        self.line.set_value(description)
        self.line.set_value(adctd)

        if self.icon:
            self.__icon_line()
        else:
            self.__polygon_line()

        self.line.set_value(data)
        self.line.set_value(morada)
        self.line.set_value(local)
        self.line.set_value(nome)
        self.line.set_value(self.azimute)
        self.line.set_value(tec)
        self.line.set_value(tipo)
        self.line.set_value(direcao)
        self.line.set_value(produto)

    def __icon_line(self):
        '''
        (lat, lon, name(hora paragon), description(cgi),
        appeddatacolumnstodescription, morada, local, nome, azimute, tec,
        tipo, direcção, produto (nº sessão))
        '''
        icon = self.icon.get_n_icon()
        iconcolor = self.icon.get_color()
        iconscale = self.icon.get_scale()
        iconheading = (self.azimute + 180) % 360
        linestringcolor = self.icon.get_line_string()

        self.line.set_value(icon)
        self.line.set_value(iconcolor)
        self.line.set_value(iconscale)
        self.line.set_value(iconheading)
        if linestringcolor != None:
            self.line.set_value(linestringcolor)

    def __polygon_line(self):
        '''
        (lat, lon, name(hora paragon), description(cgi),
        appenddatacolumnstodescription, polygon, polygoncolor, polygonazimute,
        polygonaltitude, data, morada, local, nome, azimute, tecnologia, tipo,
        direcção, produto)
        '''
        polygon = self.polygon.get_radius()
        polygoncolor = self.polygon.get_color()
        polygonazi = (self.azimute + 180) % 360
        polygonalt = self.polygon.get_altitude()
        polygonampl = self.polygon.get_amplitude()

        self.line.set_value(polygon)
        self.line.set_value(polygoncolor)
        self.line.set_value(polygonazi)
        self.line.set_value(polygonalt)
        self.line.set_value(polygonampl)

    def get_line(self):
        return self.line.get_line()


class ExcelLine(object):

    def __init__(self):
        self.line = []

    def set_value(self, value):
        self.line.append(value)

    def get_line(self):
        return self.line


class InfoConstruction(object):

    def __init__(self):
        self.file_name = None
        self.color = None

    def set_file_name(self, file_name):
        self.file_name = file_name

    def get_file_name(self):
        return self.file_name

    def set_color(self, color_name):
        self.color = color_name

    def get_color(self):
        return self.color


class Icon(InfoConstruction):

    def __init__(self):
        super(Icon, self).__init__()
        self.n_icon = 166  # icon por defeito do Google Earth
        self.scale = 1.0  # escala por defeito
        self.heading = 0  # heading por defeito
        self.line_string = ''  # se tiver um nome de cor é para desenhar

    def set_n_icon(self, n_icon):
        self.n_icon = n_icon

    def get_n_icon(self):
        return self.n_icon

    def set_scale(self, scale):
        self.scale = scale

    def get_scale(self):
        return self.scale

    def set_heading(self, heading):
        self.heading = heading

    def get_heading(self):
        return self.heading

    def set_line_string(self, color_name=''):
        self.line_string = color_name

    def get_line_string(self):
        return self.line_string


class Polygon(InfoConstruction):

    def __init__(self):
        super(Polygon, self).__init__()
        self.amplitude = None
        self.altitude = None
        self.radius = None

    def set_amplitude(self, amplitude):
        self.amplitude = amplitude

    def get_amplitude(self):
        return self.amplitude

    def set_altitude(self, altitude):
        self.altitude = altitude

    def get_altitude(self):
        return self.altitude

    def set_radius(self, radius):
        self.radius = radius

    def get_radius(self):
        return self.radius


class InfoParagon(object):

    def __init__(self):
        self.info = XlsReader()

    def open_xls(self, file_name):
        self.info.open_xls(file_name)

    def get_worksheet_names(self):
        return self.info.get_worksheet_names()

    def get_workbook(self):
        return self.info.get_workbook()

    def open_worksheet(self, worksheet_name):
        self.info.open_worksheet(worksheet_name)

    def get_n_rows(self):
        return self.info.num_rows

    def reset_counter(self):
        self.info.reset_l_conter()

    def get_next_line(self):
        return self.info.get_next_line()


class NoResults(object):

    def __init__(self):
        self.line = ExcelLine()

    def set_line(self, paragon_line, final_line):
        '''
        paragon_line - [tipo, direcção, nº produto, hora, dia, cgi]
        final_line - [lat, lon, hora, cgi, adctd, ...(icon or polygon
                        related values)..., data, morada, local, nome, azimute,
                        tecnologia, tipo, direcção, produto]

        Constructs a line:
        ['CGI', 'Data' ,Hora', 'Tipo', 'Direcção', 'Produto', 'CGI anterior',
        'Data', 'Hora', 'Nome', 'Azimute', 'Coordenadas']
        '''
        if final_line:
            coords = str(final_line[0]) + ', ' + str(final_line[1])
        else:
            final_line = ['sem dados' for x in range(20)]
            coords = 'sem dados'

        line = [paragon_line[5], paragon_line[4], paragon_line[3],
                paragon_line[0], paragon_line[1], paragon_line[2],
                final_line[3], final_line[-10], final_line[2], final_line[-7],
                final_line[-6], coords]

        for value in line:
            self.line.set_value(value)

    def get_line(self):
        return self.line.get_line()


class PostNoResults(NoResults):

    def __init__(self):
        super(PostNoResults, self).__init__()

    def set_line(self, final_line):
        '''
        final_line - [lat, lon, hora, cgi, adctd, ...(icon or polygon
                        related values)..., data, morada, local, nome, azimute,
                        tecnologia, tipo, direcção, produto, coordenadas]

        Constructs a line:
        ['CGI Posterior', 'Data', 'Hora', 'Nome', 'Azimute', 'Coordenadas']
        '''
        coords = str(final_line[0]) + ', ' + str(final_line[1])
        line = [final_line[3], final_line[-10], final_line[2], final_line[-7],
                final_line[-6], coords]

        for value in line:
            self.line.set_value(value)


class CgiTimeFilter(object):

    def __init__(self, xls, time):
        self.xls_name = xls
        self.xls_original = InfoParagon()
        self.xls_filtered = XlsWriter()
        self.time = time * 60
        self.line = None
        self.last_row = None

    def filter(self):
        self.xls_original.open_xls(self.xls_name)
        worksheet_names = self.xls_original.get_worksheet_names()
        self.xls_original.workbook = self.xls_original.get_workbook()

        for sheet_name in worksheet_names:
            self.xls_original.worksheet = self.xls_original.workbook.sheet_by_name(sheet_name)
            self.xls_original.open_worksheet(sheet_name)
            num_rows = self.xls_original.get_n_rows()
            if num_rows > 0:
                self.xls_filtered.add_sheet(sheet_name)
                self.__filter_sheet(sheet_name)

        temp = self.xls_name[:self.xls_name.rfind('.')] + '_temp.xlsx'
        self.xls_filtered.workbook.save(temp)       
        self.__clean_temp()

    def __filter_sheet(self, sheet_name):
        num_rows = self.xls_original.get_n_rows()  # total number of rows
        self.xls_original.num_cells = self.xls_original.worksheet.ncols - 1  # cols total
        stack = []  # added 13FEV15

        # Manipulate the Excel Lines ------------------------------------------
        n_row = 0
        for line in range(num_rows + 1):
            print(n_row)
            self.line = self.xls_original.get_next_line()  # raw paragon line
            if n_row == 0:  # write the firs line (titles)
                self.__write_excel(self.line)
            # append the first line (data) to the stack
            elif not self.last_row or not stack:
                stack.append(self.line)  # added 13FEV15
                num_rows += 1
            # altered 13FEV15 ----------------------------
            elif self.last_row[3] != self.line[3]:
                if self.line[2] < stack[-1][2]:
                    self.__write_excel(self.line)
                elif self.line[2] > stack[-1][2]:
                    self.__write_excel(stack.pop())
                    stack.append(self.line)                  
            #  -------------------------------------------
            else:
                if self.__calc_sec_diference(self.last_row[2], 
                                             self.line[2]) > self.time:
                    # altered 13FEV15 -----------------------------------------
                    if self.line[2] < stack[-1][2]:
                        self.__write_excel(self.line)
                    elif self.line[2] > stack[-1][2]:
                        self.__write_excel(stack.pop())
                        stack.append(self.line)                       
                    # ---------------------------------------------------------
            self.last_row = self.line
            n_row +=1
        if stack:  # added 20FEV15
            self.__write_excel(stack.pop())


    def __write_excel(self, line):
        self.xls_filtered.write_line(line)

    def __calc_sec_diference(self, time_1, time_2):
        '''
        str, str -> int
        Transforms two datetime values from the paragon excel file in python
        datetime objects, the calculate the difference between the two and
        return that diference in seconds.
        time_1 and time_2 are two strings with datetime values
        '''
        fmt = '%H:%M:%S'  # hora:minutos:segundos
        datetime_1 = datetime.strptime(time_1, fmt)
        datetime_2 = datetime.strptime(time_2, fmt)

        if datetime_2 > datetime_1:
            dif = datetime_2 - datetime_1
        else:
            return 0 #datetime_1 - datetime_2  # altered 16FEV15

        return dif.seconds

    def __clean_temp(self):
        original_path = os.getcwd()

        xls = self.xls_name[self.xls_name.rfind('/') + 1:]
        if xls[-1] == 's':
            xls += 'x'

        temp = xls[:xls.rfind('.')] + '_temp.xlsx'

        try:
            os.chdir(self.xls_name[:self.xls_name.rfind('/')])
        except:
            xls_name = self.xls_name + 'x'
            os.chdir(xls_name[:xls_name.rfind('/')])
            
        os.system('del "' + xls + '"')
        os.system('ren "' + temp + '" "' + xls + '"')
        os.chdir(original_path)


#==============================================================================
# CLASSES FROM DATA PACKAGE
#==============================================================================
class XlsReader(object):

    def __init__(self):
        self.workbook = None
        self.l_counter = -1
        self.worksheet = None
        self.num_rows = None
        self.num_cells = None

    def reset_l_conter(self):
        self.l_counter -= 1

    def open_xls(self, file_name):
        try:
            self.workbook = xlrd.open_workbook(file_name)
        except:
            return "Error"
        return self.get_worksheet_names()

    def get_workbook(self):
        return self.workbook

    def get_worksheet_names(self):
        return [x for x in self.workbook.sheet_names()]  # sheet names

    def open_worksheet(self, worksheet_name):
        self.l_counter = -1
        # sheet.Sheet()
        self.worksheet = self.workbook.sheet_by_name(worksheet_name)
        self.num_rows = self.worksheet.nrows - 1  # total number of rows
        self.num_cells = self.worksheet.ncols - 1  # total number of columns

    def get_n_rows(self):
        return self.num_rows

    def get_next_line(self):
        data_row = []
        self.l_counter += 1
        curr_cell = - 1

        while curr_cell < self.num_cells:
            curr_cell += 1
            cell_value = self.worksheet.cell_value(self.l_counter, curr_cell)
            # Format cell values to append, in Excel all ints are floats .0 ---
            if type(cell_value) is float and str(
                    cell_value)[-2:] == ".0" and \
                    self.worksheet.cell_type(self.l_counter, curr_cell) != 3:
                data_row.append(int(cell_value))
            # datetime -----------------------------------------
            elif self.worksheet.cell_type(self.l_counter, curr_cell) == 3:
                data_row.append(
                    xlrd.xldate_as_tuple(abs(cell_value), 0))
            elif type(cell_value) is float:  # real floats -----
                data_row.append(cell_value)
            else:  # unicode -----------------------------------
                try:  # In some files all values are unicode, so
                    if str(cell_value[-2] == ".0"):
                        data_row.append(int(cell_value))
                    else:
                        data_row.append(float(cell_value))
                except:
                    data_row.append(cell_value)

        return data_row


class XlsWriter(object):

    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = None
        self.sheet_name = None
        self.row_counter = 1

    def save_workbook(self, file_name, _type):
        '''
        Save the excel file builted in the same path as the original file
        type - is a str containing the type of file: icon, polygon or no_res
        '''
        out_file_name = self.__choose_file_name(file_name, _type)
        self.workbook.save(out_file_name)
        return out_file_name

    def add_sheet(self, sheet_name):
        '''
        Add a sheet to the workbook
        '''
        if self.sheet_name == None:
            self.sheet_name = sheet_name
            self.worksheet = self.workbook.active
        else:
            self.sheet_name = sheet_name
            self.worksheet = self.workbook.create_sheet()
        self.worksheet.title = str(self.sheet_name)
        self.row_counter = 1

    def write_line(self, line, col=1):
        '''
        writes a line of excel
        '''
        for i in range(len(line)):
            cell = self.worksheet.cell(row=self.row_counter, column=i + col)
            cell.value = line[i]
        self.row_counter += 1

    def __choose_file_name(self, file_name, _type):
        '''
        Builds a new output file name, based on the original file_name plus
        version number
        '''
        # directory path, no file ---------------------------------------------
        original_dir_path = os.path.abspath(os.path.dirname(file_name))
        # file_name path without extension ------------------------------------
        abs_original_file = os.path.abspath(os.path.splitext(file_name)[0])
        # file_name without extension -----------------------------------------
        original_file = abs_original_file[abs_original_file.rfind('\\') + 1:]

        new_name = ''

        if _type == 'icon':
            first = '_icons-0.1.xlsx'
            other = '_icons-'
            i_1 = -15  # for index when building the files list
            i_2 = -14
        elif _type == 'polygon':
            first = '_polygons-0.1.xlsx'
            other = '_polygons-'
            i_1 = -18
            i_2 = -17
        else:
            first = '_no_results-0.1.xlsx'
            other = '_no_results-'
            i_1 = -20
            i_2 = -19

        xlsxs = [x for x in os.listdir(original_dir_path) if
                 (x[-4:] == 'xlsx' or x[-3:] == 'xls') and
                 (x[:i_1] == original_file or x[:i_2] == original_file or
                  x[:-5] == original_file or x[:-4] == original_file)]
        xlsxs.sort()

        if not os.path.isfile(file_name[:file_name.rfind('.')] + first):
            new_name = file_name[:file_name.rfind('.')] + first
        else:
            ver = str(round(float(xlsxs[-1][-8:-5]) + .1, 2))
            new_name = file_name[:file_name.rfind('.')] + other + ver + '.xlsx'

        return new_name


class Db(object):
    '''
    5 Methods:
    __init__
    connect
    update
    export FALTA, talvez não necessário
    close
    '''

    def __init__(self):
        self.db = None
        self.cur = None

    def connect(self):
        #path = os.getcwd()[:os.getcwd().rfind(os.sep)] + os.sep + 'data'
        path = os.getcwd() + os.sep + 'data'
        db = path + os.sep + 'cgi.db'
        self.db = sqlite3.connect(db)
        self.cur = self.db.cursor()

    def update(self, excel_file):
        '''
        TODO: Isto está feito para o xls_reader do xls2kmz. Alterar p o
        xls_reader deste Paragon GIS Analyst.
        '''
        # [[[1st sheet name][1st row data][2nd row data][...][Nth row data]]]
        new_data = []

        for x in XlsReader(excel_file).read_excel()[0]:
            x.append(date.today())
            new_data.append(tuple(x))

        new_data.pop(0)  # new_data[0] - the leave name
        new_data.pop(0)  # new_data[1] - titles row
        # new_data[2:] - value rows

        self.cur.executemany('''INSERT INTO cgi(cgi, latitude, longitude,
            morada, local, nome, cp, azimute, tecnologia, date) VALUES(?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?)''', new_data)  # falta um IF DOESN'T EXIST
        self.db.commit()

    def export(self):
        pass

    def close(self):
        self.db.close()


class Query(object):
    '''
    7 Methods:
    __init__
    to_dual_tuple
    to_triple_tuple
    to_quad_tuple
    to_all_tuple
    fetch_all
    get_result
    '''

    def __init__(self, query):
        self.db = Db()
        self.db.connect()
        self.query = query
        self.result = []
        self.tuple_result = []
        self.i = 0

    def to_dual_tuple(self, query):
        while self.i < len(query):
            self.tuple_result.append((query[self.i], query[self.i + 1]))
            self.i += 2

        return self.tuple_result

    def to_triple_tuple(self, query):
        while self.i < len(query):
            self.tuple_result.append((query[self.i], query[self.i + 1],
                                      query[self.i + 2]))
            self.i += 3

        return self.tuple_result

    def to_quad_tuple(self, query):
        while self.i < len(query):
            self.tuple_result.append((query[self.i], query[self.i + 1],
                                      query[self.i + 2], query[self.i + 3]))
            self.i += 4

        return self.tuple_result

    def to_all_tuple(self, query):
        while self.i < len(query):
            self.tuple_result.append((query[self.i], query[self.i + 1],
                query[self.i + 2], query[self.i + 3], query[self.i + 4]))
            self.i += 5

        return self.tuple_result

    def fetch_all(self):
        self.result.append(self.db.cur.fetchall())

    def get_result(self):
        self.db.close()
        return self.result


class SingleQuery(Query):
    '''
    5 Methods:
    __init__
    cgi
    lat
    lon
    tec
    '''

    def __init__(self, query):
        super(SingleQuery, self).__init__(query)

    def cgi(self):
        # query is a list
        for x in self.query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ?', (x,))
            self.fetch_all()

        return self.get_result()

    def lat(self, query):
        for x in query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ?', (x,))
            self.fetch_all()

        return self.get_result()

    def lon(self, query):
        for x in query:
            self.db.cur.execute('SELECT * FROM cgi WHERE longitude = ?', (x,))
            self.fetch_all()

        return self.get_result()

    def tec(self, query):
        for x in query:
            self.db.cur.execute('SELECT * FROM cgi WHERE tecnologia = ?', (x,))
            self.fetch_all()

        return self.get_result()


class DualQuery(Query):
    '''
    7 Methods:
    __init__
    cgi_lat
    cgi_lon
    cgi_tec
    lat_lon
    lat_tec
    lon_tec
    '''

    def __init__(self, query):
        super(DualQuery, self).__init__(query)
        self.tuple_query = self.to_dual_tuple(self.query)

    def cgi_lat(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def cgi_lon(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                longitude = ?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def cgi_tec(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                tecnologia = ?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def lat_lon(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                longitude = ?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def lat_tec(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                tecnologia = ?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def lon_tec(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE longitude = ? AND \
                tecnologia = ?', (x, y,))
            self.fetch_all()

        return self.get_result()


class TripleQuery(Query):
    '''
    5 Methods:
    __init__
    cgi_lat_lon
    cgi_lat_tec
    cgi_lon_tec
    lat_lon_tec
    '''

    def __init__(self, query):
        super(TripleQuery, self).__init__(query)
        self.tuple_query = self.to_triple_tuple(self.query)

    def cgi_lat_lon(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND longitude = ?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def cgi_lat_tec(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND tecnologia = ?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def cgi_lon_tec(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                longitude = ? AND tecnologia = ?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def lat_lon_tec(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                longitude = ? AND tecnologia = ?', (x, y, z,))
            self.fetch_all()

        return self.get_result()


class QuadQuery(Query):
    '''
    2 Methods:
    __init__
    cgi_lat_lon_tec
    '''

    def __init__(self, query):
        super(QuadQuery, self).__init__(query)
        self.tuple_query = self.to_quad_tuple(self.query)

    def cgi_lat_lon_tec(self):
        for (x, y, z, a) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND longitude = ? AND tecnologia = ?',
                (x, y, z, a,))
            self.fetch_all()

        return self.get_result()


class DateSingleDualQuery(Query):
    '''
    6 methods:
    __init__
    date
    cgi_date
    lat_date
    lon_date
    tec_date
    '''

    def __init__(self, query, operator):
        super(DateSingleDualQuery, self).__init__(query)
        self.tuple_query = self.to_dual_tuple(self.query)
        self.operator = operator

    def date(self):
        for x in self.query:
            self.db.cur.execute('SELECT * FROM cgi WHERE date' \
                + self.operator + '?', (x,))
            self.fetch_all()

        return self.get_result()

    def cgi_date(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi= ? AND \
                date' + self.operator + '?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def lat_date(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                date' + self.operator + '?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def lon_date(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE longitude = ? AND \
                date' + self.operator + '?', (x, y,))
            self.fetch_all()

        return self.get_result()

    def tec_date(self):
        for (x, y) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE tecnologia = ? AND \
                date' + self.operator + '?', (x, y,))
            self.fetch_all()

        return self.get_result()


class DateTripleQuery(Query):
    '''
    7 methods:
    __init__
    cgi_lat_date
    cgi_lon_date
    cgi_tec_date
    lat_lon_date
    lat_tec_date # FALTA NO EXCEL
    lon_tec_date
    '''

    def __init__(self, query, operator):
        super(DateTripleQuery, self).__init__(query)
        self.tuple_query = self.to_triple_tuple(self.query)
        self.operator = operator

    def cgi_lat_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def cgi_lon_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                longitude = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def cgi_tec_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                tecnologia = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def lat_lon_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                longitude = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def lat_tec_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                tecnologia = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()

    def lon_tec_date(self):
        for (x, y, z) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE longitude = ? AND \
                tecnologia = ? AND date' + self.operator + '?', (x, y, z,))
            self.fetch_all()

        return self.get_result()


class DateQuadQuery(Query):
    '''
    5 Methods:
    __init__
    cgi_lat_lon_date
    cgi_lat_tec_date # FALTA NO EXCEL
    cgi_lon_tec_date # FALTA NO EXCEL
    lat_lon_tec_date
    '''

    def __init__(self, query, operator):
        super(DateQuadQuery, self).__init__(query)
        self.tuple_query = self.to_quad_tuple(self.query)
        self.operator = operator

    def cgi_lat_lon_date(self):
        for (x, y, z, a) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND longitude = ? AND \
                date' + self.operator + '?', (x, y, z, a,))
            self.fetch_all()

        return self.get_result()

    def cgi_lat_tec_date(self):
        for (x, y, z, a) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND tecnologia = ? AND \
                date' + self.operator + '?', (x, y, z, a,))
            self.fetch_all()

        return self.get_result()

    def cgi_lon_tec_date(self):
        for (x, y, z, a) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                longitude = ? AND tecnologia = ? AND \
                date' + self.operator + '?', (x, y, z, a,))
            self.fetch_all()

        return self.get_result()

    def lat_lon_tec_date(self):
        for (x, y, z, a) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE latitude = ? AND \
                longitude = ? AND tecnologia = ? AND \
                date' + self.operator + '?', (x, y, z, a,))
            self.fetch_all()

        return self.get_result()


class AllQuery(Query):

    def __init__(self, query, operator):
        super(AllQuery, self).__init__(query)
        self.tuple_query = self.to_all_tuple(self.query)
        self.operator = operator

    def cgi_lat_lon_tec_date(self):
        for (x, y, z, a, b) in self.tuple_query:
            self.db.cur.execute('SELECT * FROM cgi WHERE cgi = ? AND \
                latitude = ? AND longitude = ? AND tecnologia = ? AND \
                date' + self.operator + '?', (x, y, z, a, b,))
            self.fetch_all()

        return self.get_result()


#==============================================================================
# MAIN FUNCTION
#==============================================================================
def main():
    icon_color = None
    icon_scale = 1.0
    cgi_time = 0
    window()

#==============================================================================
# AUTO STARTER
#==============================================================================
if __name__ == '__main__':
    main()
